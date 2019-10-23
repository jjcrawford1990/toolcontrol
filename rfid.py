#module to take RFID data and output a useful RFID number and authenticate
import openpyxl

class RFID:
    #buffered_message = ['0', '02', '6B', '19', '5D', '1', '\n'] #store the buffer
    full_message = [] #read the message
    rfid_presented = [] #store only the uid of rfid
    uid_database = '/home/pi/Desktop/toolcontrol-master/uiddatabase.xlsx'  # path of the UID database file

    def __init__(self, buffered_message): #raw_data is passed as 'buffered_message' argument
        self.buffered_message = buffered_message
        self.authentication_level = 0 #default to 0 (no access)
        for i in self.buffered_message:  # this will be the ser.read() function
            # buffered_message.append(i)
            if i == 10:  # ASCII character 10, linebreak, end of message
                self.full_message = self.buffered_message
                self.buffered_message = []  # clear list
                self.message_opener(self.full_message)

    def message_opener(self, read):
        temp_str = str(read[0]) #store first char as a string
        #opener = ''.join(map(bin, bytearray(temp_str, 'ascii')))  # join to an empty string '' and assign to opener
        #opener_id = int(opener[2:])

        if temp_str == '17': #17 is first message, this could be a signifier or cal access or non cal tool access
            #continue to definition for recording all char's
            print('This is an opening message')
            self.message_comprehension(self.full_message)
        else:
            #message sent without the '0' opener, ignore?
            print('Error')

    def message_comprehension(self, full_message):
        if len(full_message) == 7: #4byte RFID UID, # start message, cage number and \n end message
            for i in full_message[1:5]: #iterate over elements 1 thru 9 of full_message list
                self.rfid_presented.append("{:02X}".format(i)) #append to empty rfid list the 4 byte UID
        else:
            print("This is not an RFID message")
        self.l_to_st = ''.join(self.rfid_presented) #string of rfid
        self.check_valid_uid()

    def check_valid_uid(self):
        #open pyxl excel
        #search in row for rfid
        #if found, set the authenticate parameter to 1
        # GPIO high for lock open & log time and rfid uid which was granted in seperate workbook
        #if not found, do nothing, perhaps flash a red LED in a seperate function
        uid_wb = openpyxl.load_workbook(self.uid_database)  # create an object of the active workbook
        uid_active_wb = uid_wb.active  # create an object of the active workbook
        uiddata_ws = uid_wb.worksheets[0]  # create an object of the first worksheet
        row_iteration = 1 #define a variable to store the row number we are at
        self.uid_found = 0 #create attribute, if not created, will throw attribute error if evaluating as UID not found.
        for row in uid_active_wb.iter_rows(min_col = 1, max_col = 1): #search only column 1
            for k in row:
                if k.value == self.l_to_st:
                    print("I found your UID", uid_active_wb.cell(row_iteration, 2).value)
                    self.uid_found_row = row_iteration #assign the row in which UID was found to this attribute
                    self.uid_found = 1
                    break
            row_iteration += 1 #increment the row we are at by 1
        if self.uid_found == 1:
            self.authentication_level = self.uid_found_row
            if self.authentication_level >= int(self.full_message[5]):
                self.authentication_level = 1
            else:
                self.authentication_level = 0
        else:
            self.authentication_level = 0